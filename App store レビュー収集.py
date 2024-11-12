import requests
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl import Workbook

# 設定変数
app_id = '763377066'  # 取得したいアプリのapp_idに変更する
excel_file_name = 'output.xlsx'  # 保存したいExcelファイル名に変更する

# 新しいExcelワークブックを作成する
wb = Workbook()
ws = wb.active
ws.title = 'シート1'

# iTunesのレビューを取得する
def get_reviews(app_id):
    entries = []
    for i in range(1, 11):  # 1ページから10ページまで取得
        # sortBy=mostRecentで最新順に取得
        xml_url = f"https://itunes.apple.com/jp/rss/customerreviews/id={app_id}/sortBy=mostRecent/page={i}/xml"
        response = requests.get(xml_url)
        xml = ET.fromstring(response.content)

        for entry in xml.findall('{http://www.w3.org/2005/Atom}entry'):
            entries.append(entry)
    
    return entries

# Excelファイルに書き込む
def write_to_excel(entries, ws):
    # ヘッダー行の設定
    headers = ["更新日時", "ID", "タイトル", "内容", "投票合計", "投票数", "評価", "バージョン", "作者名", "作者URL"]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col).value = header

    row = 2  # データの書き込み開始行を2行目に設定
    for entry in entries:
        ws.cell(row, 1).value = entry.find('{http://www.w3.org/2005/Atom}updated').text # 更新日時
        ws.cell(row, 2).value = entry.find('{http://www.w3.org/2005/Atom}id').text # ID
        ws.cell(row, 3).value = entry.find('{http://www.w3.org/2005/Atom}title').text # タイトル
        ws.cell(row, 4).value = entry.find('{http://www.w3.org/2005/Atom}content').text # 内容
        
        ws.cell(row, 5).value = entry.find('{http://itunes.apple.com/rss}voteSum').text # 投票合計
        ws.cell(row, 6).value = entry.find('{http://itunes.apple.com/rss}voteCount').text # 投票数
        ws.cell(row, 7).value = entry.find('{http://itunes.apple.com/rss}rating').text  # 評価
        ws.cell(row, 8).value = entry.find('{http://itunes.apple.com/rss}version').text # バージョン
        
        author = entry.find('{http://www.w3.org/2005/Atom}author') # 作者
        ws.cell(row, 9).value = author.find('{http://www.w3.org/2005/Atom}name').text # 作者名
        ws.cell(row, 10).value = author.find('{http://www.w3.org/2005/Atom}uri').text # 作者URL

        row += 1

entries = get_reviews(app_id)
write_to_excel(entries, ws)

# Excelファイルを保存する
wb.save(excel_file_name)